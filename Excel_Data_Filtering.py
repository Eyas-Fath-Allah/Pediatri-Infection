from openpyxl import load_workbook, Workbook

# Load the Excel file and create new xlsx file
workbook = load_workbook('C:\\Users\\eyas4\\Desktop\\Excel-İnfection\\ateş1.xlsx')
sheet = workbook.active
wbresult = Workbook()
wsresult = wbresult.active

# Get the maximum row in the sheet
max_row = sheet.max_row

x = 2
# Print the headers
for i in range(1, 21):
    wsresult.cell(row=1, column=i).value = sheet.cell(row=1, column=i).value

# Iterate over the rows, starting from the second row (assuming headers are in the first row)
for row in range(2, max_row + 1):
    age = sheet.cell(row=row, column=20).value
    if age <= 90:
        for i in range(1, 21):
            wsresult.cell(row=x, column=i).value = sheet.cell(row=row, column=i).value
        x += 1

# Save the modified Excel file
wbresult.save('C:\\Users\\eyas4\\Desktop\\Excel-İnfection\\ateş2.xlsx')
