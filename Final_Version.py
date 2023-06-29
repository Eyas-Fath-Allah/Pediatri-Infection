from openpyxl import load_workbook
from datetime import datetime

folder_path = 'C:\\Users\\eyas4\\Desktop\\Excel-İnfection'
excel_file_name = 'Test'
# Load the Excel file
workbook = load_workbook(f'{folder_path}\\{excel_file_name}.xlsx')
sheet = workbook.active

# Get the maximum row in the sheet
max_row = sheet.max_row

# Print 'Yaş(Gün)'
sheet.cell(row=1, column=20).value = 'Yaş(Gün)'

# Iterate over the rows, starting from the second row (assuming headers are in the first row)
for row in range(2, max_row + 1):
    # Get the date of birth and current date values from the respective columns
    dob = sheet.cell(row=row, column=19).value
    current_date = sheet.cell(row=row, column=1).value[6:]

    # Convert the values to datetime objects
    dob = datetime.strptime(dob, '%d/%m/%Y')  # Adjust the date format as per your Excel file
    current_date = datetime.strptime(current_date, '%d/%m/%Y')  # Adjust the date format as per your Excel file

    # Calculate the age in days
    age_in_days = (current_date - dob).days

    # Write the age in days
    sheet.cell(row=row, column=20).value = age_in_days

# Save the modified Excel file
workbook.save(f'{folder_path}\\{excel_file_name}-1.xlsx')

x = 2
for row in range(2, max_row + 1):
    age = sheet.cell(row=x, column=20).value
    if age > 90:
        sheet.delete_rows(x)
    else:
        x += 1

# Save the modified Excel file
workbook.save(f'{folder_path}\\{excel_file_name}-2.xlsx')
