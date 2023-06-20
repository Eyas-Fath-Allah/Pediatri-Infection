from openpyxl import load_workbook
from datetime import datetime

# Load the Excel file
workbook = load_workbook('C:\\Users\\eyas4\\Desktop\\Excel-İnfection\\ateş.xlsx')
sheet = workbook.active

# Get the maximum row in the sheet
max_row = sheet.max_row

# Print 'Yaş(Gün)'
sheet.cell(row=1, column=20).value = 'Yaş(Gün)'

# Iterate over the rows, starting from the second row (assuming headers are in the first row)
for row in range(2, max_row + 1):
    # Get the date of birth and current date values from the respective columns
    dob = sheet.cell(row=row, column=19).value
    current_date = sheet.cell(row=row, column=1).value[6:]

    # Convert the values to datetime objects
    dob = datetime.strptime(dob, '%d/%m/%Y')  # Adjust the date format as per your Excel file
    current_date = datetime.strptime(current_date, '%d/%m/%Y')  # Adjust the date format as per your Excel file

    # Calculate the age in days
    age_in_days = (current_date - dob).days

    # Write the age in days to the third column
    sheet.cell(row=row, column=20).value = age_in_days

# Save the modified Excel file
workbook.save('C:\\Users\\eyas4\\Desktop\\Excel-İnfection\\ateş1.xlsx')
